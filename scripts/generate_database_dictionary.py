"""生成采购智能助手数据库字段字典 Excel 和完整建表 DDL。

这个脚本把“数据库设计说明”集中在一份易读的 Python 数据结构里，再生成两个交付物：

1. ``docs/database/procurement_assistant_data_dictionary.xlsx``：给产品、架构和开发人员查看；
2. ``docs/database/procurement_assistant_schema.sql``：把 001 业务表迁移和 002 Checkpoint
   迁移拼成一份完整 DDL，方便新环境一次性审阅和执行。

脚本不会连接数据库，也不会执行 DDL。运行前需要临时安装 openpyxl：

    uv run --no-project --with openpyxl python scripts/generate_database_dictionary.py
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as error:  # pragma: no cover - 只在缺少生成工具时执行
    raise SystemExit(
        "缺少 openpyxl，请执行："
        "uv run --no-project --with openpyxl python scripts/generate_database_dictionary.py"
    ) from error


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MIGRATION_DIR = PROJECT_ROOT / "backend" / "migrations"
OUTPUT_DIR = PROJECT_ROOT / "docs" / "database"
EXCEL_PATH = OUTPUT_DIR / "procurement_assistant_data_dictionary.xlsx"
SCHEMA_PATH = OUTPUT_DIR / "procurement_assistant_schema.sql"


@dataclass(frozen=True)
class FieldSpec:
    """一个数据库字段及其面向开发者的说明。"""

    name: str
    data_type: str
    nullable: bool
    default: str
    primary_key: bool
    relation: str
    purpose: str
    example: str
    index_or_constraint: str
    sensitivity: str


@dataclass(frozen=True)
class TableSpec:
    """一张表的用途和字段列表。"""

    name: str
    category: str
    purpose: str
    authority: str
    writers: str
    readers: str
    lifecycle: str
    ddl_source: str
    notes: str
    fields: tuple[FieldSpec, ...]


@dataclass(frozen=True)
class IndexSpec:
    """一个索引或主键/检查约束的说明。"""

    table: str
    name: str
    kind: str
    columns_or_rule: str
    purpose: str
    source: str


def field(
    name: str,
    data_type: str,
    purpose: str,
    *,
    nullable: bool = False,
    default: str = "无",
    primary_key: bool = False,
    relation: str = "无物理外键；由 Delegate 做归属校验",
    example: str = "—",
    index_or_constraint: str = "—",
    sensitivity: str = "内部标识",
) -> FieldSpec:
    """用较少的重复参数创建字段说明，默认值与约束仍在调用处明确写出。"""

    return FieldSpec(
        name=name,
        data_type=data_type,
        nullable=nullable,
        default=default,
        primary_key=primary_key,
        relation=relation,
        purpose=purpose,
        example=example,
        index_or_constraint=index_or_constraint,
        sensitivity=sensitivity,
    )


def build_table_specs() -> tuple[TableSpec, ...]:
    """返回 9 张业务表和 4 张 LangGraph Checkpoint 表的完整字段字典。"""

    return (
        TableSpec(
            name="assistant_threads",
            category="业务表",
            purpose="保存前端会话归属及当前活动场景指针。",
            authority="会话归属和活动场景入口的权威记录。",
            writers="DatabaseDelegate.begin_run/get_or_create_thread/start_scenario/update_scenario_status",
            readers="Agent API、Snapshot API、Application、场景切换协调器",
            lifecycle="首次请求创建；持续保留；活动指针在场景终态清空。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="同一 thread 由 thread_id 串行；user_id 不匹配时按无权访问处理。",
            fields=(
                field("thread_id", "VARCHAR(80)", "前端会话唯一标识，也是 Checkpoint 的逻辑 thread_id。", primary_key=True, example="thread_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "会话所属用户，用于每次读写的归属校验。", example="user_10001", index_or_constraint="idx_assistant_threads_user_updated 的前导列", sensitivity="个人标识"),
                field("active_scenario_instance_id", "VARCHAR(80)", "当前会话正在运行或等待的场景实例；无活动场景时为空。", nullable=True, relation="逻辑指向 scenario_instances.scenario_instance_id", example="scenario_01J...", sensitivity="内部标识"),
                field("created_at", "TIMESTAMP WITH TIME ZONE", "会话第一次登记时间。", example="2026-08-07 10:00:00+08", sensitivity="运行元数据"),
                field("updated_at", "TIMESTAMP WITH TIME ZONE", "会话最近一次状态、Run 或场景指针变更时间。", example="2026-08-07 10:03:12+08", index_or_constraint="idx_assistant_threads_user_updated 的第二列", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="scenario_instances",
            category="业务表",
            purpose="保存一个可以跨多个 Turn 暂停、恢复和结束的 DAG 场景实例。",
            authority="场景生命周期、过期时间和当前等待组的权威记录。",
            writers="GraphRunner、Application、DatabaseDelegate.start/update/expire",
            readers="Application、SceneSwitch、Snapshot API、管理过期脚本",
            lifecycle="创建后 running/waiting；最终为 completed/aborted/expired；记录不删除。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="同一 thread 最多一个非终态活动场景由短事务和行锁保证。",
            fields=(
                field("scenario_instance_id", "VARCHAR(80)", "一次场景实例的唯一标识；同一 thread 的不同场景使用不同命名空间。", primary_key=True, example="scenario_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "场景所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="idx_scenario_thread_status 的第一列", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "场景所属用户，冗余保存以便查询和归属校验。", example="user_10001", sensitivity="个人标识"),
                field("scenario_id", "VARCHAR(100)", "静态 Scenario Registry 中的场景编号，例如 smart_routing。", example="smart_routing", index_or_constraint="应用静态目录二次校验", sensitivity="业务元数据"),
                field("input_source", "VARCHAR(32)", "场景入口来源：按钮或自然语言。", example="button", index_or_constraint="CHECK input_source", sensitivity="业务元数据"),
                field("status", "VARCHAR(32)", "场景当前生命周期状态。", example="waiting", index_or_constraint="CHECK status；idx_scenario_thread_status/idx_scenario_expiry_status", sensitivity="运行元数据"),
                field("started_at", "TIMESTAMP WITH TIME ZONE", "场景实例开始时间。", example="2026-08-07 10:00:01+08", sensitivity="运行元数据"),
                field("updated_at", "TIMESTAMP WITH TIME ZONE", "场景最近一次状态或等待组变更时间。", example="2026-08-07 10:01:20+08", sensitivity="运行元数据"),
                field("expires_at", "TIMESTAMP WITH TIME ZONE", "Checkpoint/场景等待点失效时间，默认设计为创建后 24 小时。", example="2026-08-08 10:00:01+08", index_or_constraint="idx_scenario_expiry_status 的第一列", sensitivity="运行元数据"),
                field("current_wait_group_id", "VARCHAR(80)", "当前允许用户提交的等待组编号；运行中或终态可为空。", nullable=True, relation="逻辑指向 pending_actions.wait_group_id", example="wait_01J...", index_or_constraint="与 pending_actions 事务校验", sensitivity="内部标识"),
                field("ended_at", "TIMESTAMP WITH TIME ZONE", "场景进入终态的时间。", nullable=True, example="2026-08-07 10:05:00+08", sensitivity="运行元数据"),
                field("end_reason", "VARCHAR(100)", "终止原因，例如 user_completed、deployment_expired。", nullable=True, example="user_completed", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="assistant_runs",
            category="业务表",
            purpose="登记每一次 HTTP/AG-UI Run，提供幂等、状态审计和 Trace 关联。",
            authority="一次 Run 是否重复、成功、失败或被拒绝的权威记录。",
            writers="Application/DatabaseDelegate.begin_run/finish_run/bind_run_to_scenario",
            readers="Agent API、错误处理、审计查询、Trace 查询",
            lifecycle="创建为 running；收尾为 succeeded/failed/rejected；历史长期保留。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="run_id 是全局幂等键；重复请求优先返回已有 Run 结果。",
            fields=(
                field("run_id", "VARCHAR(80)", "一次用户提交、按钮或表单恢复的唯一请求编号。", primary_key=True, example="run_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "Run 所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="idx_runs_thread_started 的第一列", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "Run 所属用户，防止跨用户复用 run_id。", example="user_10001", sensitivity="个人标识"),
                field("trace_id", "VARCHAR(80)", "本次调用链根编号，用于关联 trace_spans。", relation="逻辑关联 trace_spans.trace_id", example="trace_01J...", index_or_constraint="idx_runs_trace", sensitivity="内部标识"),
                field("input_type", "VARCHAR(64)", "本次 Run 的输入类型：natural_language、scenario_trigger、action、form_submit。", example="form_submit", sensitivity="业务元数据"),
                field("scenario_instance_id", "VARCHAR(80)", "Run 处理的场景实例；入口路由 Run 在场景创建后补写。", nullable=True, relation="逻辑指向 scenario_instances.scenario_instance_id", example="scenario_01J...", sensitivity="内部标识"),
                field("status", "VARCHAR(32)", "Run 的持久化结果状态。", example="succeeded", index_or_constraint="CHECK status", sensitivity="运行元数据"),
                field("started_at", "TIMESTAMP WITH TIME ZONE", "Run 开始时间。", example="2026-08-07 10:00:01+08", index_or_constraint="idx_runs_thread_started 的第二列", sensitivity="运行元数据"),
                field("finished_at", "TIMESTAMP WITH TIME ZONE", "Run 收尾时间；运行中为空。", nullable=True, example="2026-08-07 10:00:03+08", sensitivity="运行元数据"),
                field("error_code", "VARCHAR(100)", "失败、拒绝或重试需要展示/查询的稳定错误码。", nullable=True, example="THREAD_BUSY", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="thread_execution_leases",
            category="并发控制表",
            purpose="保存同一 thread 当前在途 Run 的短租约，保证同一会话串行。",
            authority="当前 thread 是否被某个 Run 占用。",
            writers="DatabaseDelegate.begin_run/finish_run/租约过期治理",
            readers="Run 入口准入事务",
            lifecycle="Run 接受时创建/续期；Run 收尾或租约过期时删除。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="表中通常只有短暂记录，不承担历史审计。",
            fields=(
                field("thread_id", "VARCHAR(80)", "被占用的会话唯一标识。", primary_key=True, relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("run_id", "VARCHAR(80)", "当前持有租约的 Run。", relation="逻辑指向 assistant_runs.run_id", example="run_01J...", sensitivity="内部标识"),
                field("expires_at", "TIMESTAMP WITH TIME ZONE", "租约失效时间，防止进程异常后永久占用。", example="2026-08-07 10:00:31+08", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="pending_actions",
            category="交互表",
            purpose="保存服务端签发的一次性按钮、表单、栏目选择和重试操作。",
            authority="当前用户可提交的 Action 及其输入 Schema/等待组。",
            writers="DatabaseDelegate.save_wait_request/begin_run/update_scenario_status",
            readers="Agent API 预校验、准入事务、Snapshot 投影",
            lifecycle="创建为 pending；提交后 consumed；场景结束或过期后 invalidated/expired；不删除。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="最终消费必须在短事务中 FOR UPDATE 复查，不能只信任前端提交。",
            fields=(
                field("action_id", "VARCHAR(80)", "一次性操作的唯一编号，前端提交时原样带回。", primary_key=True, example="action_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("wait_group_id", "VARCHAR(80)", "同一个等待点的一组互斥 Action 编号。", relation="逻辑关联 scenario_instances.current_wait_group_id", example="wait_01J...", index_or_constraint="idx_actions_wait_group", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "操作所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="idx_actions_thread_status_expiry 的第一列", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "操作所属用户。", example="user_10001", sensitivity="个人标识"),
                field("scenario_instance_id", "VARCHAR(80)", "操作所属场景，保证旧场景 Action 不能操作新场景。", relation="逻辑指向 scenario_instances.scenario_instance_id", example="scenario_01J...", sensitivity="内部标识"),
                field("kind", "VARCHAR(64)", "操作类型，例如 form_submit、next_page、go_custom_purchase。", example="form_submit", sensitivity="业务元数据"),
                field("input_schema_id", "VARCHAR(100)", "后端静态 Registry 中用于校验输入的 Schema 编号。", example="smart_routing.purchase_fields", sensitivity="业务元数据"),
                field("status", "VARCHAR(32)", "操作的消费状态。", example="pending", index_or_constraint="CHECK status；idx_actions_thread_status_expiry 的第二列", sensitivity="运行元数据"),
                field("payload_json", "JSONB", "服务端签发的候选项、字段定义、分页信息等安全 UI 数据。", default="'{}'::JSONB", example='{"page":1}', index_or_constraint="NOT NULL", sensitivity="业务数据"),
                field("created_at", "TIMESTAMP WITH TIME ZONE", "Action 签发时间。", example="2026-08-07 10:01:00+08", sensitivity="运行元数据"),
                field("expires_at", "TIMESTAMP WITH TIME ZONE", "Action 可提交的最后时间，通常与 Checkpoint TTL 一致。", example="2026-08-08 10:01:00+08", index_or_constraint="idx_actions_thread_status_expiry 的第三列", sensitivity="运行元数据"),
                field("consumed_at", "TIMESTAMP WITH TIME ZONE", "Action 成功消费时间。", nullable=True, example="2026-08-07 10:02:00+08", sensitivity="运行元数据"),
                field("consumed_by_run_id", "VARCHAR(80)", "实际消费该 Action 的 Run。", nullable=True, relation="逻辑指向 assistant_runs.run_id", example="run_01J...", sensitivity="内部标识"),
            ),
        ),
        TableSpec(
            name="assistant_messages",
            category="展示数据表",
            purpose="保存可恢复的用户消息和助手可展示文字。",
            authority="会话消息历史；只保存允许展示给用户的文字，不保存隐藏思维链。",
            writers="Application._persist_display_output/DatabaseDelegate.append_message",
            readers="Snapshot API、长期记忆更新、审计查询",
            lifecycle="消息写入后长期保留；不做归档和删除。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="UI 结构化块不放在本表，而放在 assistant_ui_blocks。",
            fields=(
                field("message_id", "VARCHAR(100)", "消息唯一编号；前端 AG-UI 文字消息使用同一编号。", primary_key=True, example="message_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "消息所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="idx_messages_thread_created 的第一列", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "消息所属用户，用于读权限校验。", example="user_10001", sensitivity="个人标识"),
                field("run_id", "VARCHAR(80)", "产生消息的 Run。", relation="逻辑指向 assistant_runs.run_id", example="run_01J...", sensitivity="内部标识"),
                field("role", "VARCHAR(32)", "消息角色，仅允许 user 或 assistant。", example="assistant", index_or_constraint="CHECK role", sensitivity="业务元数据"),
                field("content", "TEXT", "原始用户输入或可展示的助手文字。", example="请补充采购用途", sensitivity="业务内容"),
                field("created_at", "TIMESTAMP WITH TIME ZONE", "消息创建时间。", example="2026-08-07 10:01:02+08", index_or_constraint="idx_messages_thread_created 的第二列", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="assistant_ui_blocks",
            category="展示数据表",
            purpose="保存已经通过协议模型校验的结构化 UI CUSTOM 事件。",
            authority="页面刷新时可投影的历史 UI 块；不是 LangGraph 原始 State。",
            writers="Application/DatabaseDelegate.append_ui_block",
            readers="Snapshot API、页面恢复投影",
            lifecycle="事件写入后长期保留；快照只选择仍有展示价值的块。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="旧 Action 不因历史 UI 块直接恢复，是否可交互由当前 Checkpoint 和 Action 状态决定。",
            fields=(
                field("block_id", "BIGSERIAL", "UI 块自增顺序编号，用于同一 thread 的稳定排序。", primary_key=True, example="10001", index_or_constraint="PRIMARY KEY；idx_ui_blocks_thread_id 的第二列", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "UI 块所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="idx_ui_blocks_thread_id 的第一列", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "UI 块所属用户。", example="user_10001", sensitivity="个人标识"),
                field("block_json", "JSONB", "经过 Pydantic 校验的 AG-UI CUSTOM 事件完整 JSON。", example='{"type":"CUSTOM","name":"procurement.form"}', index_or_constraint="NOT NULL", sensitivity="业务数据"),
                field("created_at", "TIMESTAMP WITH TIME ZONE", "UI 块写入时间。", example="2026-08-07 10:01:02+08", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="user_memories",
            category="个性化表",
            purpose="保存每个用户自动生成和合并的长期记忆 JSON。",
            authority="个人长期记忆；只用于非关键个性化表达和推荐辅助。",
            writers="MemoryUpdater/DatabaseDelegate.merge_memory",
            readers="MemoryUpdater、允许读取记忆的模型任务",
            lifecycle="按用户持续更新；当前设计长期保留，不提供用户查看/更正/删除入口。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="关键采购分支不得依赖 memory_json；更新使用短事务锁定最新 JSON 合并补丁。",
            fields=(
                field("user_id", "VARCHAR(128)", "用户唯一标识，同时作为记忆主键。", primary_key=True, example="user_10001", index_or_constraint="PRIMARY KEY", sensitivity="个人标识"),
                field("memory_json", "JSONB", "系统自动维护的个人记忆对象。", default="'{}'::JSONB", example='{"preferred_language":"zh-CN"}', index_or_constraint="NOT NULL", sensitivity="个性化业务数据"),
                field("source_run_id", "VARCHAR(80)", "最近一次成功生成记忆补丁的 Run。", nullable=True, relation="逻辑指向 assistant_runs.run_id", example="run_01J...", sensitivity="内部标识"),
                field("updated_at", "TIMESTAMP WITH TIME ZONE", "记忆最近一次成功更新的时间。", example="2026-08-07 10:03:00+08", sensitivity="运行元数据"),
                field("last_error", "VARCHAR(1000)", "最近一次记忆异步更新失败的安全错误摘要；不影响业务 Run。", nullable=True, example="MODEL_TIMEOUT", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="trace_spans",
            category="可观测性表",
            purpose="保存整个 Run 的父子调用链、首包/首字/最终结果和耗时。",
            authority="调用耗时与执行状态查询数据；不反向改变业务结果。",
            writers="TraceCollector/ OpenGaussTraceDelegate.save_spans",
            readers="运维 SQL 查询、问题排查、容量分析",
            lifecycle="Span 收尾后批量写入；当前设计不归档、不删除。",
            ddl_source="backend/migrations/001_assistant_core.sql",
            notes="输入输出按当前“不脱敏持久化”决策保存，但 safe_json 会排除凭据字段。",
            fields=(
                field("span_id", "VARCHAR(80)", "当前调用步骤的唯一编号。", primary_key=True, example="span_01J...", index_or_constraint="PRIMARY KEY", sensitivity="内部标识"),
                field("trace_id", "VARCHAR(80)", "一条完整调用链的根编号。", example="trace_01J...", index_or_constraint="idx_trace_trace_started 的第一列；查询主键", sensitivity="内部标识"),
                field("parent_span_id", "VARCHAR(80)", "父步骤编号；根 Span 为空。", nullable=True, relation="逻辑自关联 trace_spans.span_id", example="span_01J...", sensitivity="内部标识"),
                field("run_id", "VARCHAR(80)", "Span 所属 Run。", relation="逻辑指向 assistant_runs.run_id", example="run_01J...", index_or_constraint="idx_trace_run", sensitivity="内部标识"),
                field("thread_id", "VARCHAR(80)", "Span 所属会话，方便按会话排查。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", sensitivity="内部标识"),
                field("user_id", "VARCHAR(128)", "Span 所属用户，方便归属查询。", example="user_10001", sensitivity="个人标识"),
                field("span_kind", "VARCHAR(32)", "Span 层级，例如 HTTP、GRAPH、MODEL、AGENT、DATABASE。", example="AGENT", index_or_constraint="应用枚举校验", sensitivity="运行元数据"),
                field("name", "VARCHAR(200)", "步骤名称，例如 POST /api/v1/agent 或 delegate.query。", example="delegate.column_recognition", index_or_constraint="idx_trace_name_started 的第一列", sensitivity="运行元数据"),
                field("target", "VARCHAR(200)", "实际调用目标的安全标识或 URL 模板名称。", nullable=True, example="column-recognition-agent", sensitivity="运行元数据"),
                field("attempt", "INTEGER", "同一逻辑调用的尝试次数，从 1 开始。", default="无（应用默认 1）", example="1", index_or_constraint="CHECK attempt >= 1", sensitivity="运行元数据"),
                field("status", "VARCHAR(32)", "Span 状态：RUNNING、OK、ERROR、CANCELLED。", example="OK", index_or_constraint="CHECK status；idx_trace_status_started 的第一列", sensitivity="运行元数据"),
                field("started_at", "TIMESTAMP WITH TIME ZONE", "Span 开始时间。", example="2026-08-07 10:01:00+08", index_or_constraint="idx_trace_trace_started 的第二列", sensitivity="运行元数据"),
                field("finished_at", "TIMESTAMP WITH TIME ZONE", "Span 结束时间；进行中为空。", nullable=True, example="2026-08-07 10:01:01+08", sensitivity="运行元数据"),
                field("duration_ms", "DOUBLE PRECISION", "从开始到结束的总耗时毫秒数。", nullable=True, example="1012.5", sensitivity="运行元数据"),
                field("first_byte_ms", "DOUBLE PRECISION", "从开始到收到外围响应首字节的耗时毫秒数。", nullable=True, example="220.4", sensitivity="运行元数据"),
                field("first_text_delta_ms", "DOUBLE PRECISION", "从开始到收到首个可展示文字增量的耗时毫秒数。", nullable=True, example="315.8", sensitivity="运行元数据"),
                field("final_result_ms", "DOUBLE PRECISION", "从开始到结构化最终结果完成的耗时毫秒数。", nullable=True, example="998.2", sensitivity="运行元数据"),
                field("input_json", "JSONB", "本步骤输入的结构化快照，便于问题排查。", nullable=True, example='{"product_name":"笔记本"}', sensitivity="业务数据"),
                field("output_json", "JSONB", "本步骤输出的结构化结果或安全摘要。", nullable=True, example='{"status":"ok"}', sensitivity="业务数据"),
                field("error_code", "VARCHAR(100)", "失败或取消时的稳定错误码。", nullable=True, example="DELEGATE_TIMEOUT", sensitivity="运行元数据"),
                field("attributes_json", "JSONB", "扩展属性，例如 delegate_id、模型名、页面来源。", default="'{}'::JSONB", example='{"delegate_id":"ioi"}', index_or_constraint="NOT NULL", sensitivity="运行元数据"),
            ),
        ),
        TableSpec(
            name="checkpoint_migrations",
            category="LangGraph Checkpoint 表",
            purpose="记录官方 Checkpointer 已执行的迁移版本。",
            authority="LangGraph Checkpointer 迁移版本状态。",
            writers="显式数据库迁移；应用启动禁止 setup() 自动修改。",
            readers="Checkpointer 版本检查和运维核对",
            lifecycle="随 Checkpointer 版本升级追加；不删除。",
            ddl_source="backend/migrations/002_langgraph_checkpoint.sql",
            notes="当前固定到 langgraph-checkpoint-postgres 3.1.1 的迁移版本 0–9。",
            fields=(
                field("v", "INTEGER", "官方 Checkpointer 迁移版本号。", primary_key=True, example="9", index_or_constraint="PRIMARY KEY", sensitivity="基础设施元数据"),
            ),
        ),
        TableSpec(
            name="checkpoints",
            category="LangGraph Checkpoint 表",
            purpose="保存 LangGraph 根 Graph 的检查点元数据和序列化状态。",
            authority="DAG interrupt/resume 的流程状态。",
            writers="LangGraph AsyncPostgresSaver",
            readers="LangGraph GraphRunner resume/list/get_state",
            lifecycle="随 Graph 执行产生；当前设计不归档、不删除。",
            ddl_source="backend/migrations/002_langgraph_checkpoint.sql",
            notes="thread_id 使用 assistant_threads.thread_id；checkpoint_ns 使用 scenario_instance_id。",
            fields=(
                field("thread_id", "TEXT", "LangGraph 配置中的会话编号。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="PRIMARY KEY 第一列；checkpoints_thread_id_idx", sensitivity="内部标识"),
                field("checkpoint_ns", "TEXT", "Checkpoint 命名空间；本项目使用场景实例编号隔离同一会话的不同场景。", relation="逻辑指向 scenario_instances.scenario_instance_id", default="''", example="scenario_01J...", index_or_constraint="PRIMARY KEY 第二列", sensitivity="内部标识"),
                field("checkpoint_id", "TEXT", "LangGraph 生成的单个检查点编号。", primary_key=False, example="1f0...", index_or_constraint="PRIMARY KEY 第三列", sensitivity="内部标识"),
                field("parent_checkpoint_id", "TEXT", "当前检查点的父检查点编号；根检查点为空。", nullable=True, example="1ef...", sensitivity="内部标识"),
                field("type", "TEXT", "检查点序列化类型标识，由官方 serde 决定。", nullable=True, example="msgpack", sensitivity="基础设施元数据"),
                field("checkpoint", "JSONB", "检查点元数据和 channel_versions 等结构化状态。", example='{"channel_versions":{}}', index_or_constraint="NOT NULL", sensitivity="业务运行状态"),
                field("metadata", "JSONB", "LangGraph 检查点附加元数据。", default="'{}'::JSONB", example='{"step":2}', index_or_constraint="NOT NULL", sensitivity="业务运行状态"),
            ),
        ),
        TableSpec(
            name="checkpoint_blobs",
            category="LangGraph Checkpoint 表",
            purpose="保存 Graph channel 的二进制序列化值，并按版本供检查点引用。",
            authority="LangGraph channel value blob。",
            writers="LangGraph AsyncPostgresSaver",
            readers="LangGraph 恢复 channel 状态",
            lifecycle="随检查点产生；当前设计不归档、不删除。",
            ddl_source="backend/migrations/002_langgraph_checkpoint.sql",
            notes="blob 在官方最终迁移中允许为空；不要自行改为 NOT NULL。",
            fields=(
                field("thread_id", "TEXT", "所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="PRIMARY KEY 第一列；checkpoint_blobs_thread_id_idx", sensitivity="内部标识"),
                field("checkpoint_ns", "TEXT", "所属场景命名空间。", relation="逻辑指向 scenario_instances.scenario_instance_id", default="''", example="scenario_01J...", index_or_constraint="PRIMARY KEY 第二列", sensitivity="内部标识"),
                field("channel", "TEXT", "Graph channel 名称。", example="messages", index_or_constraint="PRIMARY KEY 第三列", sensitivity="基础设施元数据"),
                field("version", "TEXT", "channel 的版本字符串。", example="00000000000000000001.0", index_or_constraint="PRIMARY KEY 第四列", sensitivity="基础设施元数据"),
                field("type", "TEXT", "blob 的序列化类型标识。", example="msgpack", sensitivity="基础设施元数据"),
                field("blob", "BYTEA", "channel 的二进制序列化内容。", nullable=True, example="二进制", sensitivity="业务运行状态"),
            ),
        ),
        TableSpec(
            name="checkpoint_writes",
            category="LangGraph Checkpoint 表",
            purpose="保存检查点对应的任务写入和 pending writes，用于恢复与任务语义。",
            authority="LangGraph pending writes。",
            writers="LangGraph AsyncPostgresSaver",
            readers="LangGraph 恢复任务发送和 channel 写入",
            lifecycle="随 Graph 执行产生；当前设计不归档、不删除。",
            ddl_source="backend/migrations/002_langgraph_checkpoint.sql",
            notes="task_path 是 3.1.1 最终迁移增加的字段，不能省略。",
            fields=(
                field("thread_id", "TEXT", "所属会话。", relation="逻辑指向 assistant_threads.thread_id", example="thread_01J...", index_or_constraint="PRIMARY KEY 第一列；checkpoint_writes_thread_id_idx", sensitivity="内部标识"),
                field("checkpoint_ns", "TEXT", "所属场景命名空间。", relation="逻辑指向 scenario_instances.scenario_instance_id", default="''", example="scenario_01J...", index_or_constraint="PRIMARY KEY 第二列", sensitivity="内部标识"),
                field("checkpoint_id", "TEXT", "写入所属的检查点编号。", relation="逻辑关联 checkpoints.checkpoint_id", example="1f0...", index_or_constraint="PRIMARY KEY 第三列", sensitivity="内部标识"),
                field("task_id", "TEXT", "产生写入的 LangGraph 任务编号。", example="task_01J...", index_or_constraint="PRIMARY KEY 第四列", sensitivity="基础设施元数据"),
                field("task_path", "TEXT", "任务在子图/任务树中的路径。", default="''", example="node:subgraph", sensitivity="基础设施元数据"),
                field("idx", "INTEGER", "同一任务写入的顺序编号。", example="0", index_or_constraint="PRIMARY KEY 第五列", sensitivity="基础设施元数据"),
                field("channel", "TEXT", "写入目标 channel。", example="messages", sensitivity="基础设施元数据"),
                field("type", "TEXT", "写入值的序列化类型标识。", nullable=True, example="msgpack", sensitivity="基础设施元数据"),
                field("blob", "BYTEA", "写入值的二进制序列化内容。", example="二进制", index_or_constraint="NOT NULL", sensitivity="业务运行状态"),
            ),
        ),
    )


def build_index_specs() -> tuple[IndexSpec, ...]:
    """返回迁移文件中的所有索引、主键和检查约束。"""

    rows: list[IndexSpec] = []
    for table in build_table_specs():
        primary_key = [item.name for item in table.fields if item.primary_key]
        if table.name == "checkpoints":
            primary_key = ["thread_id", "checkpoint_ns", "checkpoint_id"]
        elif table.name == "checkpoint_blobs":
            primary_key = ["thread_id", "checkpoint_ns", "channel", "version"]
        elif table.name == "checkpoint_writes":
            primary_key = ["thread_id", "checkpoint_ns", "checkpoint_id", "task_id", "idx"]
        rows.append(
            IndexSpec(
                table.name,
                f"pk_{table.name}",
                "PRIMARY KEY",
                ", ".join(primary_key),
                "保证表内记录唯一；Checkpoint 表使用官方复合主键。",
                table.ddl_source,
            )
        )

    rows.extend(
        (
            IndexSpec("assistant_threads", "idx_assistant_threads_user_updated", "INDEX", "user_id, updated_at", "按用户查询最近会话。", "001"),
            IndexSpec("scenario_instances", "idx_scenario_thread_status", "INDEX", "thread_id, status", "查询会话当前活动/终态场景。", "001"),
            IndexSpec("scenario_instances", "idx_scenario_expiry_status", "INDEX", "expires_at, status", "扫描需要惰性或批量过期的场景。", "001"),
            IndexSpec("assistant_runs", "idx_runs_thread_started", "INDEX", "thread_id, started_at", "按会话查看 Run 时间线。", "001"),
            IndexSpec("assistant_runs", "idx_runs_trace", "INDEX", "trace_id", "按 trace_id 查找 Run。", "001"),
            IndexSpec("pending_actions", "idx_actions_thread_status_expiry", "INDEX", "thread_id, status, expires_at", "查找会话当前有效/过期操作。", "001"),
            IndexSpec("pending_actions", "idx_actions_wait_group", "INDEX", "wait_group_id", "按等待组定位互斥 Action。", "001"),
            IndexSpec("assistant_messages", "idx_messages_thread_created", "INDEX", "thread_id, created_at", "按创建顺序恢复会话消息。", "001"),
            IndexSpec("assistant_ui_blocks", "idx_ui_blocks_thread_id", "INDEX", "thread_id, block_id", "按事件顺序读取会话 UI 块。", "001"),
            IndexSpec("trace_spans", "idx_trace_trace_started", "INDEX", "trace_id, started_at", "按调用链和时间查看父子 Span。", "001"),
            IndexSpec("trace_spans", "idx_trace_run", "INDEX", "run_id", "查询一次 Run 的完整 Span。", "001"),
            IndexSpec("trace_spans", "idx_trace_name_started", "INDEX", "name, started_at", "按调用步骤聚合耗时。", "001"),
            IndexSpec("trace_spans", "idx_trace_status_started", "INDEX", "status, started_at", "查询错误/取消 Span。", "001"),
            IndexSpec("checkpoints", "checkpoints_thread_id_idx", "INDEX", "thread_id", "官方 Checkpointer 按会话列出检查点。", "002"),
            IndexSpec("checkpoint_blobs", "checkpoint_blobs_thread_id_idx", "INDEX", "thread_id", "官方 Checkpointer 按会话读取 blob。", "002"),
            IndexSpec("checkpoint_writes", "checkpoint_writes_thread_id_idx", "INDEX", "thread_id", "官方 Checkpointer 按会话读取 pending writes。", "002"),
            IndexSpec("scenario_instances", "ck_scenario_input_source", "CHECK", "input_source IN ('button', 'natural_language')", "限制场景入口来源。", "001"),
            IndexSpec("scenario_instances", "ck_scenario_status", "CHECK", "status IN ('running', 'waiting', 'completed', 'aborted', 'expired')", "限制场景生命周期状态。", "001"),
            IndexSpec("assistant_runs", "ck_run_status", "CHECK", "status IN ('running', 'succeeded', 'failed', 'rejected')", "限制 Run 状态。", "001"),
            IndexSpec("pending_actions", "ck_action_status", "CHECK", "status IN ('pending', 'consumed', 'invalidated', 'expired')", "限制一次性操作状态。", "001"),
            IndexSpec("assistant_messages", "ck_message_role", "CHECK", "role IN ('user', 'assistant')", "只保存用户或可展示助手消息。", "001"),
            IndexSpec("trace_spans", "ck_trace_attempt", "CHECK", "attempt >= 1", "重试次数从 1 开始。", "001"),
            IndexSpec("trace_spans", "ck_trace_status", "CHECK", "status IN ('RUNNING', 'OK', 'ERROR', 'CANCELLED')", "限制 Span 状态。", "001"),
        )
    )
    return tuple(rows)


ENUM_ROWS: tuple[tuple[str, str, str, str, str], ...] = (
    ("scenario_instances.input_source", "input_source", "button", "页面按钮触发场景", "否"),
    ("scenario_instances.input_source", "input_source", "natural_language", "自然语言经 ReAct 选择场景", "否"),
    ("scenario_instances.status", "status", "running", "场景正在执行", "否"),
    ("scenario_instances.status", "status", "waiting", "场景暂停，等待用户提交 Action/Form", "否"),
    ("scenario_instances.status", "status", "completed", "场景正常完成", "是"),
    ("scenario_instances.status", "status", "aborted", "场景因错误、取消或切换而中止", "是"),
    ("scenario_instances.status", "status", "expired", "Checkpoint/部署过期，旧场景不能恢复", "是"),
    ("assistant_runs.status", "status", "running", "Run 正在执行", "否"),
    ("assistant_runs.status", "status", "succeeded", "Run 正常完成", "是"),
    ("assistant_runs.status", "status", "failed", "Run 执行失败", "是"),
    ("assistant_runs.status", "status", "rejected", "Run 在入口被拒绝，例如忙或输入非法", "是"),
    ("pending_actions.status", "status", "pending", "等待用户消费", "否"),
    ("pending_actions.status", "status", "consumed", "已被某个 Run 消费", "是"),
    ("pending_actions.status", "status", "invalidated", "被同组 Action 或场景终态使失效", "是"),
    ("pending_actions.status", "status", "expired", "超过有效期", "是"),
    ("assistant_messages.role", "role", "user", "用户输入", "否"),
    ("assistant_messages.role", "role", "assistant", "允许展示给用户的助手文字", "否"),
    ("assistant_runs.input_type", "input_type", "natural_language", "自然语言消息", "否"),
    ("assistant_runs.input_type", "input_type", "scenario_trigger", "按钮触发场景", "否"),
    ("assistant_runs.input_type", "input_type", "action", "提交一次性按钮", "否"),
    ("assistant_runs.input_type", "input_type", "form_submit", "提交结构化表单", "否"),
    ("trace_spans.span_kind", "span_kind", "HTTP", "HTTP 请求根步骤", "否"),
    ("trace_spans.span_kind", "span_kind", "REACT", "ReAct 场景路由", "否"),
    ("trace_spans.span_kind", "span_kind", "SCENARIO", "场景生命周期编排", "否"),
    ("trace_spans.span_kind", "span_kind", "GRAPH", "LangGraph 执行", "否"),
    ("trace_spans.span_kind", "span_kind", "NODE", "Graph 节点", "否"),
    ("trace_spans.span_kind", "span_kind", "MODEL", "模型调用", "否"),
    ("trace_spans.span_kind", "span_kind", "AGENT", "外围 Agent 调用", "否"),
    ("trace_spans.span_kind", "span_kind", "SERVICE", "搜索、知识、排队等外部服务", "否"),
    ("trace_spans.span_kind", "span_kind", "DATABASE", "数据库 Delegate 调用", "否"),
    ("trace_spans.span_kind", "span_kind", "MEMORY", "长期记忆更新", "否"),
    ("trace_spans.status", "status", "RUNNING", "Span 执行中", "否"),
    ("trace_spans.status", "status", "OK", "Span 成功", "是"),
    ("trace_spans.status", "status", "ERROR", "Span 失败", "是"),
    ("trace_spans.status", "status", "CANCELLED", "Span 被取消", "是"),
)


RELATION_ROWS: tuple[tuple[str, str, str, str, str, str], ...] = (
    ("assistant_threads", "active_scenario_instance_id", "scenario_instances", "scenario_instance_id", "逻辑指针", "当前会话活动场景；不建物理外键以保留状态变更事务弹性。"),
    ("scenario_instances", "thread_id", "assistant_threads", "thread_id", "逻辑归属", "场景归属会话。"),
    ("scenario_instances", "current_wait_group_id", "pending_actions", "wait_group_id", "逻辑等待组", "当前等待组内的 Action 必须可提交。"),
    ("assistant_runs", "thread_id", "assistant_threads", "thread_id", "逻辑归属", "Run 归属会话。"),
    ("assistant_runs", "scenario_instance_id", "scenario_instances", "scenario_instance_id", "逻辑归属", "Run 处理的场景实例。"),
    ("assistant_runs", "trace_id", "trace_spans", "trace_id", "调用链关联", "Run 与父子 Span 关联。"),
    ("thread_execution_leases", "run_id", "assistant_runs", "run_id", "占用者", "租约由一个在途 Run 持有。"),
    ("pending_actions", "scenario_instance_id", "scenario_instances", "scenario_instance_id", "逻辑归属", "旧场景 Action 不得操作新场景。"),
    ("pending_actions", "consumed_by_run_id", "assistant_runs", "run_id", "消费记录", "记录实际消费 Action 的 Run。"),
    ("assistant_messages", "run_id", "assistant_runs", "run_id", "来源关联", "消息由某次 Run 产生。"),
    ("assistant_ui_blocks", "thread_id", "assistant_threads", "thread_id", "逻辑归属", "UI 块属于会话。"),
    ("user_memories", "source_run_id", "assistant_runs", "run_id", "来源关联", "记忆补丁来自最近成功 Run。"),
    ("trace_spans", "parent_span_id", "trace_spans", "span_id", "父子 Span", "根 Span 为空，其余 Span 组成调用树。"),
    ("trace_spans", "run_id", "assistant_runs", "run_id", "调用关联", "Span 属于某次 Run。"),
    ("checkpoints", "thread_id", "assistant_threads", "thread_id", "Checkpointer 逻辑归属", "LangGraph thread_id 与前端会话一致。"),
    ("checkpoints", "checkpoint_ns", "scenario_instances", "scenario_instance_id", "Checkpointer 命名空间", "同一会话的不同场景隔离。"),
    ("checkpoint_blobs", "thread_id/checkpoint_ns", "checkpoints", "thread_id/checkpoint_ns", "Checkpointer 复合归属", "channel blob 由检查点命名空间引用。"),
    ("checkpoint_writes", "thread_id/checkpoint_ns/checkpoint_id", "checkpoints", "thread_id/checkpoint_ns/checkpoint_id", "Checkpointer 写入关联", "pending writes 属于某个检查点。"),
)


def _iter_fields(tables: Iterable[TableSpec]) -> Iterable[tuple[TableSpec, FieldSpec]]:
    for table in tables:
        for item in table.fields:
            yield table, item


def validate_migrations(tables: tuple[TableSpec, ...]) -> None:
    """确认字段字典没有与实际 SQL 漂移。

    这里不仅检查“字典中的内容能否在 SQL 找到”，还会检查 SQL 中是否出现了
    字典没有登记的表或字段。这样以后有人直接修改迁移文件时，生成命令会立即
    失败，而不是悄悄生成一份不完整的 Excel。
    """

    sql_by_source = {
        "001": (MIGRATION_DIR / "001_assistant_core.sql").read_text(encoding="utf-8"),
        "002": (MIGRATION_DIR / "002_langgraph_checkpoint.sql").read_text(encoding="utf-8"),
    }
    expected_by_source: dict[str, set[str]] = {"001": set(), "002": set()}
    for table in tables:
        expected_by_source[table.ddl_source.rsplit("/", 1)[-1][:3]].add(table.name)

    for source, source_sql in sql_by_source.items():
        actual_tables = set(re.findall(r"CREATE TABLE IF NOT EXISTS\s+(\w+)", source_sql))
        expected_tables = expected_by_source[source]
        if actual_tables != expected_tables:
            missing = sorted(expected_tables - actual_tables)
            extra = sorted(actual_tables - expected_tables)
            raise RuntimeError(
                f"{source} 迁移表集合与字段字典不一致；缺少：{missing}；多出：{extra}"
            )

    for table in tables:
        source_sql = sql_by_source[table.ddl_source.rsplit("/", 1)[-1][:3]]
        match = re.search(
            rf"CREATE TABLE IF NOT EXISTS\s+{re.escape(table.name)}\s*\((.*?)\n\);",
            source_sql,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if match is None:
            raise RuntimeError(f"DDL 中找不到表：{table.name}")
        body = match.group(1)
        actual_fields = {
            field_name
            for field_name in re.findall(r"^\s*([A-Za-z_]\w*)\s+", body, flags=re.MULTILINE)
            if field_name.upper() not in {"PRIMARY", "CHECK", "CONSTRAINT", "UNIQUE", "FOREIGN"}
        }
        expected_fields = {item.name for item in table.fields}
        if actual_fields != expected_fields:
            missing = sorted(expected_fields - actual_fields)
            extra = sorted(actual_fields - expected_fields)
            raise RuntimeError(
                f"{table.name} 字段集合与字段字典不一致；缺少：{missing}；多出：{extra}"
            )
        for item in table.fields:
            if re.search(rf"^\s*{re.escape(item.name)}\s+", body, flags=re.MULTILINE) is None:
                raise RuntimeError(f"DDL 中找不到字段：{table.name}.{item.name}")


def _style_sheet(sheet, headers: tuple[str, ...], rows: list[tuple[object, ...]]) -> None:
    """给工作表添加筛选、冻结、换行和适合中文阅读的列宽。"""

    header_fill = PatternFill("solid", fgColor="5277C3")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1EE")
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFF")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for column_index in range(1, len(headers) + 1):
        values = [str(sheet.cell(row=row, column=column_index).value or "") for row in range(1, sheet.max_row + 1)]
        width = min(max(max(len(value) for value in values) + 2, 12), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def build_workbook(tables: tuple[TableSpec, ...]) -> Workbook:
    """创建包含表、字段、约束、枚举和关系的可筛选工作簿。"""

    workbook = Workbook()
    workbook.remove(workbook.active)

    table_headers = ("序号", "表名", "分类", "表作用", "权威内容", "主要写入方", "主要读取方", "生命周期", "DDL 来源", "设计备注")
    table_rows = [
        (index, table.name, table.category, table.purpose, table.authority, table.writers, table.readers, table.lifecycle, table.ddl_source, table.notes)
        for index, table in enumerate(tables, start=1)
    ]
    sheet = workbook.create_sheet("表清单")
    _style_sheet(sheet, table_headers, table_rows)

    field_headers = ("序号", "表名", "字段名", "数据类型", "必填", "默认值", "主键", "逻辑关联", "字段作用", "示例/取值", "索引或约束", "数据敏感性")
    field_rows = []
    for index, (table, item) in enumerate(_iter_fields(tables), start=1):
        field_rows.append((index, table.name, item.name, item.data_type, "否" if item.nullable else "是", item.default, "是" if item.primary_key else "否", item.relation, item.purpose, item.example, item.index_or_constraint, item.sensitivity))
    sheet = workbook.create_sheet("字段字典")
    _style_sheet(sheet, field_headers, field_rows)

    index_headers = ("序号", "表名", "名称", "类型", "字段或规则", "用途", "来源")
    index_rows = [(index, row.table, row.name, row.kind, row.columns_or_rule, row.purpose, row.source) for index, row in enumerate(build_index_specs(), start=1)]
    sheet = workbook.create_sheet("索引与约束")
    _style_sheet(sheet, index_headers, index_rows)

    enum_headers = ("数据位置", "字段", "值", "含义", "是否终态")
    sheet = workbook.create_sheet("状态枚举")
    _style_sheet(sheet, enum_headers, list(ENUM_ROWS))

    relation_headers = ("源表", "源字段", "目标表", "目标字段", "关联类型", "用途")
    sheet = workbook.create_sheet("逻辑关系")
    _style_sheet(sheet, relation_headers, list(RELATION_ROWS))

    execution_headers = ("顺序", "交付物/步骤", "说明", "是否可直接执行")
    execution_rows = [
        (1, "backend/migrations/001_assistant_core.sql", "创建 9 张业务表及其索引/检查约束。", "是（需先验证目标 OpenGauss 类型兼容性）"),
        (2, "backend/migrations/002_langgraph_checkpoint.sql", "创建 4 张 LangGraph Checkpoint 表，版本基线为 langgraph-checkpoint-postgres 3.1.1。", "是（需先完成 OpenGauss 兼容性验收）"),
        (3, "运行应用健康检查", "确认 assistant_* 与 checkpoint_* 13 张表都存在；应用启动不自动建表。", "否（由部署/运维执行）"),
        (4, "生产变更管理", "迁移账号执行 DDL；应用账号仅授予 DML 和必要序列权限；执行前备份。", "否（需按公司流程）"),
    ]
    sheet = workbook.create_sheet("执行说明")
    _style_sheet(sheet, execution_headers, execution_rows)

    workbook.properties.title = "采购智能助手数据库字段字典"
    workbook.properties.subject = "13 张业务与 LangGraph Checkpoint 表"
    workbook.properties.creator = "Procurement Assistant"
    workbook.properties.description = (
        "生成日期："
        f"{datetime.now(timezone.utc).date().isoformat()}；"
        "由 scripts/generate_database_dictionary.py 生成。"
    )
    return workbook


def write_schema_reference() -> None:
    """把两份有序迁移拼成一份完整的参考 DDL。"""

    header = """-- 采购智能助手完整数据库建表 DDL（参考合并版）\n--\n-- 执行顺序等同于：001_assistant_core.sql -> 002_langgraph_checkpoint.sql。\n-- 本文件便于一次性审阅/初始化新环境；生产变更仍建议按两份迁移文件逐步执行。\n-- 当前目标 OpenGauss 兼容性尚未完成生产验收，执行前必须验证 JSONB、BYTEA、复合主键、\n-- ON CONFLICT、带时区时间、LangGraph interrupt/resume 和并发行为。\n\n"""
    content = header
    for filename in ("001_assistant_core.sql", "002_langgraph_checkpoint.sql"):
        content += f"-- ===== BEGIN {filename} =====\n\n"
        content += (MIGRATION_DIR / filename).read_text(encoding="utf-8").rstrip()
        content += f"\n\n-- ===== END {filename} =====\n\n"
    # 迁移之间保留空行，但整个文件只保留一个标准结尾换行，便于 Git 严格检查。
    SCHEMA_PATH.write_text(content.rstrip() + "\n", encoding="utf-8")


def main() -> None:
    tables = build_table_specs()
    validate_migrations(tables)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_schema_reference()
    workbook = build_workbook(tables)
    workbook.save(EXCEL_PATH)
    print(f"已生成：{EXCEL_PATH}")
    print(f"已生成：{SCHEMA_PATH}")
    print(f"表数量：{len(tables)}；字段数量：{sum(len(table.fields) for table in tables)}")


if __name__ == "__main__":
    main()
